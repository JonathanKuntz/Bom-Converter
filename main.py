import csv
from ntpath import join
import sys
import os.path
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
from openpyxl.styles import Alignment
import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter import messagebox
from datetime import date


# Git Test
filename = ""       #initialisation for filname variable
excelHeaderProjektDate = date.today()
#User Interface opens the Windows explorer
root = tk.Tk()
root.geometry('400x400')
root.title("Bom Converter")


#Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
def Askfilename():
    global filename
    filename = askopenfilename() # show an "Open" dialog box and return the path to the selected file

    entry1.delete(0, "end")
    entry1.insert(0, filename)

    return filename

def return_entry(en):

    global content
    content = en.get()

    #return content


def CsvReader(filename):
    newOrderWithoutR1 = [0, 4, 18, 16, 10, 11, 12, 3, 5, 17, 13, 14, 15, 9]
    newOrderWithR1 = [0, 4, 20, 16, 10, 11, 12, 18, 19, 3, 5, 17, 13, 14, 15, 9]

    global R1Check 
    R1Check = False
    newOrder = newOrderWithoutR1
    thtCounter = 0
    smdCounter = 0
    completeListCounter = 0
    kompletteListe = []
    smdListe = []
    thtListe = []

    #Read the Bom CSV
    with open(filename, encoding='utf-8')  as csv_datei:
        reader = csv.reader(csv_datei, delimiter=';' )

        header = next(reader)   #delete first line
        if header[18] == "R1":
            newOrder = newOrderWithR1
            R1Check = True

    # geht jede Zeile der geöffneten CSV Datei duch und orded nach der der Order in der nächsten for schleife
        for zeile in reader:
            csvReaderListe = []

            for spalte in newOrder:
                data = zeile[spalte]
                csvReaderListe.append(data)
            completeListCounter += 1
            hilfsListe = [completeListCounter] + csvReaderListe
            kompletteListe.append(hilfsListe)
    #auswahlt oh die Zeile ein THT oder SMD bauteil ist
            if csvReaderListe[13] == 'THT':
                thtCounter += 1
                csvReaderListe = [thtCounter] + csvReaderListe
                thtListe.append(csvReaderListe)

            else:
                smdCounter += 1
                csvReaderListe = [smdCounter] + csvReaderListe
                smdListe.append(csvReaderListe)


        #       0               1       2           3                   4           5
    return kompletteListe, smdListe, thtListe, completeListCounter, smdCounter, thtCounter



def CreateExcelFiles(listToConvert, listToConvertCounter, excelHeaderProjektHinweis):

    #sortiert die liste nach dem BR_Value, es fehlt, dass die erste Spalte nicht mitsortiert wird
    #sort_order = ['n.b.', ' ']
    listToConvert.sort(key=lambda row:(row[4], -row[7]))

    columnNameWithoutR1 = ['Pos.', 'Menge', 'Name', 'TEC-Artikel-Nr.:', 'Wert', 'Wert 2', 'Wert 3', 'Wert 4', 'Bauform',
                            'Beschreibung', 'Hersteller', 'Lieferant 1', 'Lieferant 2', 'Briechle Artikel', 'Bauart']
    columnNameWithR1 = ['Pos.', 'Menge', 'Name', 'TEC-Artikel-Nr.:', 'Wert', 'Wert 2', 'Wert 3', 'Wert 4', 'R1', 'R2', 'Bauform',
                            'Beschreibung', 'Hersteller', 'Lieferant 1', 'Lieferant 2', 'Briechle Artikel', 'Bauart']


    #anzahlSpalten = len(columnName)    # gibt die anzahl der Werte in columnName an

    # initialize openpyxl and set sheet one to active to write on
    wb = Workbook()
    sheet = wb.active

    # Header Values
    excelHeaderProjektDatei = excelHeaderProjektName + '_STL-' + excelHeaderProjektHinweis + '_' + excelHeaderProjektVersion + '.xlsx'

    # appending Datas to sheet
    # Excel Sheet Header
    sheet["A1"] = "Projekt:"
    sheet["A2"] = "Version:"
    sheet["A3"] = "Datum:"
    sheet["A4"] = "Hinweis:"
    sheet["A5"] = "Datei:"
    sheet["A6"] = ""
    sheet["B1"] = excelHeaderProjektName
    sheet["B2"] = excelHeaderProjektVersion
    sheet["B3"] = excelHeaderProjektDate
    sheet["B4"] = excelHeaderProjektHinweis + " Stückliste"
    sheet["B5"] = excelHeaderProjektDatei

    #checks for the right list type
    columnName = columnNameWithoutR1
    
    widthLetter = 'O'
        # change column width
    sheet.column_dimensions['B'].width = 7
    sheet.column_dimensions['C'].width = 32
    sheet.column_dimensions['D'].width = 14
    sheet.column_dimensions['E'].width = 16
    sheet.column_dimensions['F'].width = 16
    sheet.column_dimensions['G'].width = 16
    sheet.column_dimensions['H'].width = 16
    sheet.column_dimensions['I'].width = 16
    sheet.column_dimensions['J'].width = 36
    sheet.column_dimensions['K'].width = 36
    sheet.column_dimensions['L'].width = 30
    sheet.column_dimensions['M'].width = 30
    sheet.column_dimensions['N'].width = 14

    if R1Check :
        columnName = columnNameWithR1
        widthLetter = 'Q'

        # change column width
        sheet.column_dimensions['B'].width = 7
        sheet.column_dimensions['C'].width = 32
        sheet.column_dimensions['D'].width = 14
        sheet.column_dimensions['E'].width = 16
        sheet.column_dimensions['F'].width = 16
        sheet.column_dimensions['G'].width = 16
        sheet.column_dimensions['H'].width = 16
        sheet.column_dimensions['I'].width = 9
        sheet.column_dimensions['J'].width = 9
        sheet.column_dimensions['K'].width = 16
        sheet.column_dimensions['L'].width = 36
        sheet.column_dimensions['M'].width = 36
        sheet.column_dimensions['N'].width = 30
        sheet.column_dimensions['O'].width = 30
        sheet.column_dimensions['P'].width = 14
    
    sheet.append(columnName)

    # append datas from List to sheet one by obe
    for item in listToConvert:
        sheet.append(item)

    # Font properties
    def set_border(ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def set_Bold(ws, cell_range):
        for row in ws[cell_range]:
            for cell in row:
                cell.font = Font(bold=True)

    # creats Borders for the datas
    lengthOfList = listToConvertCounter + 7
    set_border(sheet, 'A7:'+ widthLetter + str(lengthOfList))
    set_Bold(sheet, 'A1:B5')
    set_Bold(sheet, 'A7:' + widthLetter + '7')

    # Farbcode:8DB4E2
    # Farbcode 2: ACB9CA
    redFill = PatternFill(start_color='ACB9CA',
                          end_color='ACB9CA',
                          fill_type='solid')

    for cell in sheet["D7:D" + str(lengthOfList)]:
        cell[0].fill = redFill
    # sheet["D7"].font  = Font(color = "94A4BA")

    # Wraptext für Spalte
    def set_wrapText(ws, cell_range):
        for row in ws[cell_range]:
            for cell in row:
                cell.alignment = Alignment(wrapText=True)

    set_wrapText(sheet, 'C7:C' + str(lengthOfList))


    # abspeichern der Datei in dem selben Path wo die Datei her kommt
    saveFile = os.path.join(os.path.dirname(filename), excelHeaderProjektDatei)
    #print(savefile)
    wb.save(filename=saveFile)

    return excelHeaderProjektDatei


def createTextFile(liste, listenNamen):
    #Form wie die ERP Import datei aus zu sehen hat:
    #------------->Artikelnummer    Hierachie    Art    Anzahl    Stueckartikel    Bemerkung
    # Für uns:----> BR-Artikel          Pos.      ??    Menge       Wert            ??
    #in Spalte:--->       15            0                  4

    OrderWithotR1 = [13, 0, 0, 1, 0, 0]
    OrderWithR1 = [1, 1, 1, 1, 1]

    NewOrder = OrderWithotR1
    if R1Check:
        NewOrder = OrderWithR1


    # Filename to write #Name von der aus xlsx erstellten namen nehmen und mit txt joinen
    txtFilename = listenNamen + ".txt"

    #to create File in same Folder as the original File
    txtFilename = os.path.join(os.path.dirname(filename), txtFilename)
    # Open the file with writing permission
    temp_file = open(txtFilename, 'w')

    header = ['Artikelnummer', 'Hierachie', 'Art', 'Anzahl', 'Stueckartikel', 'Bemerkung']
    temp_file.write("%s\n" % header)

    for item in liste:
        temp_file.write("%s\n" % item)





    # Close the file
    temp_file.close()


def execute():

    return_entry(entry1)
    global filename
    filename = content
    #print(os.path.dirname(filename))
    return_entry(entry2)
    global excelHeaderProjektName
    excelHeaderProjektName = content
    #print(excelHeaderProjektName)
    return_entry(entry3)
    global excelHeaderProjektVersion
    excelHeaderProjektVersion = content
    #print(excelHeaderProjektVersion)
    return_entry(entry4)
    global excelHeaderProjektDate
    excelHeaderProjektDate = content
    #print(excelHeaderProjektDate)

    #checkt mit der splitext dem hinteren Teil ob es eine csv Datei ist
    if os.path.splitext(filename)[1] !='.csv':
        tk.messagebox.showerror("!Error!", "Bitte eine .CSV Datei auswählen oder eigeben "
                                           "\nder Komplette Pfad muss mit angegeben werden")

    else:
        CsvReader(filename)
        completeFileName = ""
        smdFileName = ""
        thtFileName = ""

        #complete
        if cB1var.get() == 1:
            CreateExcelFiles(CsvReader(filename)[0], CsvReader(filename)[3], "Komplette")
            completeFileName = CreateExcelFiles(CsvReader(filename)[0], CsvReader(filename)[3], "Komplette")
        #SMD
        if cB2var.get() == 1:
            CreateExcelFiles(CsvReader(filename)[1], CsvReader(filename)[4], "SMD")
            smdFileName = CreateExcelFiles(CsvReader(filename)[1], CsvReader(filename)[4], "SMD")
        #THT
        if cB3var.get() == 1:
            CreateExcelFiles(CsvReader(filename)[2], CsvReader(filename)[5], "THT")
            thtFileName = CreateExcelFiles(CsvReader(filename)[2], CsvReader(filename)[5], "THT")
        #ERP
        if cB4var.get() == 1:
            createTextFile(CsvReader(filename)[0], "ERP-Stücklistenimport")


        #final output message
        messagebox.showinfo("Fertig :D", "folgende Dateien wurden erstellt: \n"+completeFileName+" \n"+smdFileName+"\n"
                            +thtFileName+"\n \nAm Speicherort:\n"+os.path.dirname(filename))


#Button for searching the CSV data
button1 = tk.Button(root, text="Durchsuchen", command=Askfilename).grid(row=0, column=2)
button2 = tk.Button(root, text="Fertigstellen", command=execute).grid(row=9, column=2)
button3 = tk.Button(root, text="Abbrechen", command=sys.exit).grid(row=10, column=2)

label1 = tk.Label(root, text="Datei(.csv):").grid(row=0)
label2 = tk.Label(root, text="Projektname:").grid(row=1)
label3 = tk.Label(root, text="Version:").grid(row=2)
label4 = tk.Label(root, text="Datum:").grid(row=3)

entry1 = tk.Entry(root, width=40)
entry1.grid(row=0, column=1)
entry2 = tk.Entry(root, width=40)
entry2.grid(row=1, column=1)
entry3 = tk.Entry(root, width=40)
entry3.grid(row=2, column=1)
entry4 = tk.Entry(root, width=40)
entry4.grid(row=3, column=1)
entry4.insert(10, str(excelHeaderProjektDate))

cB1var = tk.IntVar()
cB1 = tk.Checkbutton(root, text="Komplette", variable=cB1var).grid(row=6, column=1)
cB2var = tk.IntVar()
cB2var.set(True)
cB2 = tk.Checkbutton(root, text="SMD", variable=cB2var).grid(row=7, column=1)
cB3var = tk.IntVar()
cB3var.set(True)
cB3 = tk.Checkbutton(root, text="THT", variable=cB3var).grid(row=8, column=1)
cB4var = tk.IntVar()
cB4 = tk.Checkbutton(root, text="ERP", variable=cB4var).grid(row=9, column=1)

root.mainloop()

"""
________________________Comment Block___________________

#Methode zum erstellen von CSV Dateien
def CSV_Creator(csvName, uebergabeListe):
    csvName = csvName + '.csv'
    with open(csvName, 'w',encoding = 'utf-8', newline='') as newFile:
        mywriter = csv.writer(newFile, delimiter=';', dialect='excel')
        mywriter.writerows(uebergabeListe)
    print(uebergabeListe)

#Create the csv Files
CSV_Creator('thtBom', thtListe)
CSV_Creator('completeBom', kompletteListe)

---------------------------------------------------------------
txt Datei filter
 for zeile in liste:
        temp_file = []

        for spalte in NewOrder:
           data = zeile[spalte]
           temp_file.append(data)

"""